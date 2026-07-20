from openpyxl import Workbook
from openpyxl.styles import Font

wb=Workbook()

bold=Font(bold=True)

def add_key_value_sheet(name,data):

    ws=wb.create_sheet(name)

    ws.append(["Field","Value"])

    ws["A1"].font=bold
    ws["B1"].font=bold

    for k,v in data:

        ws.append([k,v])

    for column in ws.columns:

        length=max(len(str(c.value or "")) for c in column)

        ws.column_dimensions[column[0].column_letter].width=length+4

header=[

("Contract Note No","CNT-26/27-60197922"),
("Trade Date","16/07/2026"),
("Settlement No","2026132"),
("Settlement Date","17/07/2026"),
("Client ID","BC0109"),
("Client Name","PRABHUKUMAR A")

]

summary=[

("ISIN","INE036D01028"),
("Symbol","KARURVYSYA"),
("Buy Qty",100),
("Average Price",305.7570),
("Buy Value",30595.70),
("Sell Qty",0),
("Sell Value",0),
("Net Qty",100),
("Net Obligation",30595.70)

]

charges=[

("Brokerage",20.00),
("Exchange Charges",0.94),
("IGST",3.77),
("STT",31.00),
("SEBI Fees",0.03),
("Stamp Duty",5.00),
("Net Amount",30636.44)

]

wb.remove(wb.active)

add_key_value_sheet("Header",header)
add_key_value_sheet("Summary",summary)
add_key_value_sheet("Charges",charges)

ws=wb.create_sheet("Trades")

cols=[
"Order No",
"Order Time",
"Trade No",
"Trade Time",
"Symbol",
"ISIN",
"Side",
"Exchange",
"Quantity",
"Brokerage",
"Price",
"Total"
]

ws.append(cols)

for c in ws[1]:

    c.font=bold

ws.append([
"1200000022507231",
"10:21:20",
"402236600",
"10:21:20",
"KARURVYSYA",
"",
"BUY",
"NSE",
14,
2.8004,
305.80,
4281.20
])

ws.append([
"1200000022507231",
"10:21:20",
"402236599",
"10:21:20",
"KARURVYSYA",
"",
"BUY",
"NSE",
86,
17.1996,
305.75,
26294.50
])

for column in ws.columns:

    length=max(len(str(c.value or "")) for c in column)

    ws.column_dimensions[column[0].column_letter].width=length+4

wb.save("ContractNote_Output.xlsx")

print("Created ContractNote_Output.xlsx")
